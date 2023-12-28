from __future__ import annotations

from typing import Protocol, ClassVar, Any, Type

import pandas as pd
from attrs import define
from openpyxl import Workbook
from openpyxl.styles import Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

import glytrait


class Output(Protocol):
    """Manage the output written to an `openpyxl.Worsheet`.

    Class attributes:
        name: The name of the output.

    Methods:
        write: Write the output to the worksheet.
    """

    name: ClassVar[str]
    """The name of the output."""

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        """Write the output to the worksheet.

        Args:
            data (dict[str, Any]): The data to be written, with the keys being the data names,
                and the values being the data.
            ws (Worksheet): The worksheet to write to.
        """
        raise NotImplementedError


output_classes: list[Type] = []


def register_output(cls: Type[Output]) -> Type[Output]:
    """Register an output class."""
    output_classes.append(cls)
    return cls


@register_output
@define
class SummaryOutput(Output):
    """Write the summary sheet."""

    name: ClassVar[str] = "Summary"

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        config = data["config"]
        direct_triats = data["direct_traits"]
        derived_traits = data["derived_traits"]

        HEADER = "__HEADER__"
        records = [
            # option name, option Value
            ("GlyTrait version", glytrait.__version__),
            ("Options", HEADER),
            ("Input file", config.get("input_file")),
            ("Output file", config.get("output_file")),
            ("Structure file", config.get("structure_file")),
            ("Formula file", config.get("formula_file")),
            ("Mode", config.get("mode")),
            ("Database", config.get("database")),
            ("Glycan filter ratio", config.get("filter_glycan_max_na")),
            ("Imputation method", config.get("impute_method")),
            ("Normalization method", "Total abundance"),
            ("Sialic acid linkage", config.get("sia_linkage")),
            ("Post filtering", config.get("post_filtering")),
            ("Correlation threshold", config.get("corr_threshold")),
            ("Correlation method", config.get("corr_method")),
            ("Result Overview", HEADER),
            ("Num. of samples", len(direct_triats.index)),
            ("Num. of glycans", len(direct_triats.columns)),
            ("Num. of traits", len(derived_traits.columns)),
        ]

        for i, (name, value) in enumerate(records, start=1):
            cell1, cell2 = f"A{i}", f"B{i}"
            if value == HEADER:
                ws[cell1] = name
                ws.merge_cells(f"{cell1}:{cell2}")
                ws[cell1].font = ws[cell1].font.copy(bold=True)
                ws[cell1].alignment = Alignment(horizontal="center", vertical="center")
            else:
                ws[cell1] = name
                ws[cell2] = value if value is not None else "none"

        ws.column_dimensions["A"].width = max(len(cell.value) for cell in ws["A"])
        for cell in ws["A"]:
            cell.alignment = Alignment(horizontal="left")
        for cell in ws["B"]:
            cell.alignment = Alignment(horizontal="left")


def _write_values(df: pd.DataFrame, ws: Worksheet) -> None:
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for cell in ws[1][1:]:
        cell.border = cell.border.copy(bottom=Side(border_style="thin"))


@register_output
@define
class GlycanAbundOutput(Output):
    """Write the glycan abundance."""

    name: ClassVar[str] = "Glycan abundance"

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        _write_values(data["direct_traits"], ws)


@register_output
@define
class TraitValueOutput(Output):
    """Write the trait value sheet."""

    name: ClassVar[str] = "Trait values"

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        _write_values(data["derived_traits"], ws)


@register_output
@define
class TraitDefinitionsOutput(Output):
    """Write the trait defination sheet."""

    name: ClassVar[str] = "Trait definitions"

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        formulas = data["formulas"]
        ws.append(["Trait Name", "Description"])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        for formula in formulas:
            ws.append([formula.name, formula.description])


@register_output
@define
class MetaPropertiesOutput(Output):
    """Write the meta-properties sheet."""

    name: ClassVar[str] = "Meta properties"

    def write(self, data: dict[str, Any], ws: Worksheet) -> None:
        meta_prop_df = data["meta_prop_df"]
        for row in dataframe_to_rows(meta_prop_df, index=True, header=True):
            ws.append(row)
        ws.delete_rows(2)
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)


def write_output(data: dict[str, Any], file: str) -> None:
    """Write the output file.

    Args:
        data (dict[str, Any]): The data to be written, with the keys being the data names,
            and the values being the data.
        file (str): The output filepath.
    """
    wb = Workbook()
    ws = wb.active
    first_cls = output_classes[0]
    first_output = first_cls()
    first_output.write(data, ws)
    for cls in output_classes[1:]:
        output = cls()
        ws = wb.create_sheet(output.name)
        output.write(data, ws)
    wb.save(file)
