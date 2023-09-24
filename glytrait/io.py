"""Input and output functions.

Functions:
    read_input: Read the input file.
    read_group: Read the group file.
    read_structure: Read the structure file.
    load_default_structures: Load the default structure database.
    write_output: Write the output file.
"""
from __future__ import annotations

import functools
from importlib.resources import files, as_file
from pathlib import Path
from typing import NoReturn, Literal, Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

import glytrait
from glytrait.config import Config
from glytrait.exception import *
from glytrait.formula import TraitFormula
from glytrait.glycan import Structure, load_structures


def check_input_file(df: pd.DataFrame) -> NoReturn:
    """Check the input dataframe."""
    if df.columns[0] != "Composition":
        raise InputError("The first column of the input file should be Composition.")
    if df["Composition"].duplicated().sum() > 0:
        raise InputError("There are duplicated Compositions in the input file.")

    has_struc_col = "Structure" in df.columns
    if has_struc_col:
        if df.columns[1] != "Structure":
            raise InputError("The second column of the input file should be Structure.")
        if df["Structure"].duplicated().sum() > 0:
            raise InputError("There are duplicated Structures in the input file.")

    if np.any(df.iloc[:, 2 if has_struc_col else 1 :].dtypes != "float64"):
        raise InputError("The abundance columns in the input file should be numeric.")


@functools.singledispatch
def read_group_file(file) -> pd.Series:
    """Read the group file.

    The first column should be the sample IDs and the second column should be the group names.

    Args:
        file (str | Path | pandas.DataFrame): The group file.

    Returns:
        groups (pd.Series): The group names, with sample IDs as index.

    Raises:
        NotImplementedError: If the input file type is not supported.
    """
    raise NotImplementedError(f"Cannot read group file of type {type(file)}.")


def _read_group_file(df: pd.DataFrame) -> pd.Series:
    if df.shape[1] != 1:
        raise InputError("The group file should only have two columns.")
    groups = df.squeeze()
    return groups


@read_group_file.register(str)
@read_group_file.register(Path)
def _(file) -> pd.Series:
    df = pd.read_csv(file, index_col=0)
    return _read_group_file(df)


@read_group_file.register(pd.DataFrame)
def _(df) -> pd.Series:
    return _read_group_file(df)


@functools.singledispatch
def read_structure_file(file, compositions: Iterable[str]) -> list[Structure]:
    """Read the structure file.

    `file` could be either a CSV file (or DataFrame) with all structures,
    or a directory containing the GLYCOCT files of the structures.
    If it is a CSV file, the first column should be the glycan compositions,
    and the second column should be the glycan structures.
    If it is a directory, the filenames should be the glycan compositions,
    and the file contents should be the condensed glycoCT strings.
    Only the structures of the glycans in `compositions` will be read, and the order of the
    structures will be the same as that of `compositions`.

    Args:
        file (str | Path | pandas.DataFrame):
            The CSV structure file, or a directory containing the GLYCOCT files.
        compositions (Iterable[str]): The compositions of the glycans to be read.

    Returns:
        list[Structure]: The glycans.

    Raises:
        InputError:
            If the input file format is wrong,
            or some compositions are not found in the structure file.
    """
    raise NotImplementedError(f"Cannot read structure file of type {type(file)}.")


@read_structure_file.register(str)
@read_structure_file.register(Path)
def _(file, compositions: Iterable[str]) -> list[Structure]:
    if Path(file).is_dir():
        struc_strings = _read_structure_string_from_directory(file, compositions)
    else:
        struc_strings = _read_structure_string_from_csv(file, compositions)
    return load_structures(compositions, struc_strings)


@read_structure_file.register(pd.DataFrame)
def _(df, compositions: Iterable[str]) -> list[Structure]:
    struc_strings = _read_structure_string_from_df(df, compositions)
    return load_structures(compositions, struc_strings)


def _read_structure_string_from_csv(
    file: str, compositions: Iterable[str]
) -> list[str]:
    df = pd.read_csv(file, index_col=0)
    return _read_structure_string_from_df(df, compositions)


def _read_structure_string_from_df(
    df: pd.DataFrame, compositions: Iterable[str]
) -> list[str]:
    if df.shape[1] != 1:
        raise InputError("The structure file should only have two columns.")
    structures = df.squeeze()
    struc_strings = []
    for comp in compositions:
        try:
            struc_strings.append(structures[comp])
        except KeyError:
            raise InputError(f"{comp} is not found in the structure file.")
    return struc_strings


def _read_structure_string_from_directory(
    path: str, compositions: Iterable[str]
) -> list[str]:
    path = Path(path)
    struc_strings = []
    for comp in compositions:
        try:
            filename = comp + ".glycoct_condensed"
            struc_strings.append((path / filename).read_text())
        except FileNotFoundError:
            raise InputError(f"{comp} is not provided in the structure directory.")
    return struc_strings


built_in_db = {
    "serum": files("glytrait.resources").joinpath("serum_structures.csv"),
    "IgG": files("glytrait.resources").joinpath("IgG_structures.csv"),
}


def load_default_structures(
    db: Literal["serum", "IgG"], compositions: Iterable[str]
) -> list[Structure]:
    """Return a list of default structures.

    Args:
        db: The database to use. Either "serum" or "IgG".
        compositions: The glycan compositions to search for.

    Returns:
        A list of default structures, in the order of the compositions.
    """
    db_file = built_in_db[db]
    with as_file(db_file) as db_filename:
        return read_structure_file(str(db_filename), compositions)


def write_output(
    config: Config,
    derived_traits: pd.DataFrame,
    direct_traits: pd.DataFrame,
    meta_prop_df: pd.DataFrame,
    formulas: list[TraitFormula],
) -> None:
    """Write the output file.

    Args:
        config (Config): The configuration.
        derived_traits (pd.DataFrame): The trait values, with samples as index and trait names as
            columns.
        direct_traits (pd.DataFrame): The normalized abundance table, with samples as index and
            glycan IDs as columns.
        meta_prop_df (pd.DataFrame): The table of meta properties generated by
            `build_meta_property_table`.
        formulas (list[TraitFormula]): The trait formulas.
        groups (pd.Series): The group names, with sample IDs as index. Optional.
        hypothesis_test (Optional[pd.DataFrame], optional): The hypothesis test results. Optional.
        roc_result (Optional[pd.DataFrame], optional): The ROC results. Optional.
    """
    wb = Workbook()
    ws_summary: Worksheet = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, config, direct_traits, derived_traits)

    # The trait table
    ws_trait = wb.create_sheet("Trait values")
    _write_trait_values(ws_trait, direct_traits, derived_traits)

    # The trait definitions
    ws_def = wb.create_sheet("Trait definitions")
    _write_trait_defination(ws_def, formulas)

    # The meta properties
    ws_meta = wb.create_sheet("Meta properties")
    _write_meta_properties(ws_meta, meta_prop_df)

    wb.save(config.get("output_file"))


def _write_summary(
    ws: Worksheet,
    config: Config,
    direct_triats: pd.DataFrame,
    derived_traits: pd.DataFrame,
) -> None:
    """Write the summary sheet."""
    HEADER = "__HEADER__"
    records = [
        # option name, option Value
        ("GlyTrait version", glytrait.__version__),
        ("Options", HEADER),
        ("Input file", config.get("input_file")),
        ("Output file", config.get("output_file")),
        ("Structure file", config.get("structure_file")),
        ("Formula file", config.get("formula_file")),
        ("Mode", config.get("mode")),
        ("Database", config.get("database")),
        ("Glycan filter ratio", config.get("filter_glycan_max_na")),
        ("Imputation method", config.get("impute_method")),
        ("Normalization method", "Total abundance"),
        ("Sialic acid linkage", config.get("sia_linkage")),
        ("Post filtering", config.get("post_filtering")),
        ("Correlation threshold", config.get("corr_threshold")),
        ("Correlation method", config.get("corr_method")),
        ("Result Overview", HEADER),
        ("Num. of samples", len(direct_triats.index)),
        ("Num. of glycans", len(direct_triats.columns)),
        ("Num. of traits", len(derived_traits.columns)),
    ]

    for i, (name, value) in enumerate(records, start=1):
        cell1, cell2 = f"A{i}", f"B{i}"
        if value == HEADER:
            ws[cell1] = name
            ws.merge_cells(f"{cell1}:{cell2}")
            ws[cell1].font = ws[cell1].font.copy(bold=True)
            ws[cell1].alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws[cell1] = name
            ws[cell2] = value if value is not None else "none"

    ws.column_dimensions["A"].width = max(len(cell.value) for cell in ws["A"])
    for cell in ws["A"]:
        cell.alignment = Alignment(horizontal="left")
    for cell in ws["B"]:
        cell.alignment = Alignment(horizontal="left")


def _write_trait_values(
    ws: Worksheet,
    direct_traits: pd.DataFrame,
    derived_traits: pd.DataFrame,
) -> None:
    """Write the trait values sheet."""
    combined_df = pd.concat([direct_traits, derived_traits], axis=1)

    for row in dataframe_to_rows(combined_df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1][1:]:
        cell.font = cell.font.copy(bold=True)
        cell.border = cell.border.copy(bottom=Side(border_style="double"))
    ws.insert_rows(1)
    ws.merge_cells(
        start_row=1,
        start_column=2,
        end_row=1,
        end_column=len(direct_traits.columns) + 1,
    )
    ws.cell(1, 2).value = "Direct traits"
    ws.merge_cells(
        start_row=1,
        start_column=len(direct_traits.columns) + 2,
        end_row=1,
        end_column=len(combined_df.columns) + 1,
    )
    ws.cell(1, len(direct_traits.columns) + 2).value = "Derived traits"
    for cell in ws[1][1:]:
        cell.font = cell.font.copy(bold=True)
        cell.border = cell.border.copy(bottom=Side(border_style="thin"))


def _write_trait_defination(ws: Worksheet, formulas: list[TraitFormula]) -> None:
    """Write the trait defination sheet."""
    ws.append(["Trait Name", "Description"])
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
    for formula in formulas:
        ws.append([formula.name, formula.description])


def _write_meta_properties(ws: Worksheet, meta_prop_df: pd.DataFrame) -> None:
    for row in dataframe_to_rows(meta_prop_df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
